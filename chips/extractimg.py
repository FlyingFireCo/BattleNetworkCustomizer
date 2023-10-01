import openpyxl
from PIL import Image
import os
from io import BytesIO

def save_images_from_sheet(worksheet):
    # Create images directory if doesn't exist
    if not os.path.exists('images'):
        os.makedirs('images')

    # Find the column number for "Image"
    image_col_num = None
    for col_num, col_cells in enumerate(worksheet.iter_cols()):
        if col_cells[0].value == "Image":
            image_col_num = col_num + 1  # Columns start from 1 in openpyxl
            break
    if not image_col_num:
        print(f"Error: Couldn't find the 'Image' column in sheet {worksheet.title}.")
        return
    
    # Iterate through worksheet's images
    idx = 1
    for img in worksheet._images:
        # Determine if the anchor is OneCell or TwoCell type
        if hasattr(img.anchor, 'row'):
            # For OneCellAnchor type
            row_num = img.anchor.row
            col_num = img.anchor.col
        elif hasattr(img.anchor, '_from') and hasattr(img.anchor, 'ext'):
            # For the adjusted TwoCellAnchor type
            start_pos = getattr(img.anchor, '_from')
            row_num = start_pos.row
            col_num = start_pos.col
            print(f"Image: {img}, row: {row_num}, col: {col_num}")
        else:
            print(f"Unknown anchor type for image: {img}. Attributes: {dir(img.anchor)}")
            continue

        try:
            cell = worksheet.cell(row=row_num+1, column=image_col_num)  # Adjust row number as openpyxl starts from 1
            print(f"Cell: {cell}")
            print(f"Cell value: {cell.value}")
            image_name = cell.value
        except AttributeError:
            print(f"Could not find cell for image: {img}")
            image_name = f'image{idx}'

        image_filename = f'images/{image_name}'  # Assume all images are png

        # Save the image using PIL
        with Image.open(BytesIO(img._data())) as pil_img:
            pil_img.save(image_filename)
            print(f'Saved: {image_filename}')
            idx += 1


def save_images_from_excel(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    # Process first three sheets
    sheets = workbook.sheetnames[:3]
    for sheet_name in sheets:
        print(f"Processing sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        save_images_from_sheet(worksheet)

# Replace 'chips.xlsx' with your Excel file path
save_images_from_excel('chips.xlsx')